
import csv
import re
import datetime
import pandas as pd
from openpyxl import load_workbook
import xlwings

excel_filename = "DEAS.xlsx"
# excel_filename = "EquipmentInventory.xlsx"

# def excelWriter(arcDict, sheet_name):
    # '''Writes arc dictionaries to the excel file. Not in use.'''

    # print("Writing " + sheet_name)
    # arcList = []
    # arcList.append(["Xi", "Yi", "Zi", "Xj", "Yj", "Zj", "Item", "Lij", "Uij", "Cij"])
    # for arc in arcDict.keys():
    #     arcList.append([arc[0][0], arc[0][1], arc[0][2], arc[1][0], arc[1][1],
    #         arc[1][2], arc[2], arcDict[arc][0], arcDict[arc][1], arcDict[arc][2]])

def excelWriter(df, sheet_name, xl_filename):
    book = load_workbook(xl_filename)
    sheet_names = book.get_sheet_names()
    if sheet_name in sheet_names:
        sheet = book.get_sheet_by_name(sheet_name)
        book.remove_sheet(sheet)
        book.save(xl_filename)
    writer = pd.ExcelWriter(xl_filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheet_name, index=False, index_label=False, header=True)
    writer.save()

def excelReader():
    '''Read in relevant data from Excel'''

    xl = pd.ExcelFile(excel_filename)

    # INVENTORY
    inventory_df = xl.parse("Inventory")
    inventory_rows = inventory_df.values.tolist()

    inventory_dict = {}
    total_inventory_dict = {}
    for row in inventory_rows:
        if any(row):
            room = row[0]
            commodity = row[1]
            quantity = float(row[2])
            inventory_dict[(room, commodity)] = quantity
            if commodity in total_inventory_dict:
                total_inventory_dict[commodity] += quantity
            else:
                total_inventory_dict[commodity] = quantity

    # STORAGE ROOM CAPACITY
    storage_df = xl.parse("Storage Rooms")
    storage_rows = storage_df.values.tolist()

    storage_cap_dict = {}
    for row in storage_rows:
        if any(row):
            room  = row[0]
            capacity = float(row[1]) * float(row[2])
            storage_cap_dict[room] = capacity

    # MOVEMENT COST
    df = xl.parse("Cost Data", header=None)
    rows = df.values.tolist()
    cost_dict = {}
    for rowIndex in range (1, len(rows)):
        for columnIndex in range (1, (rowIndex + 1)):
            cost_dict[(rows[rowIndex][0], rows[0][columnIndex])] = rows[rowIndex][columnIndex]
            cost_dict[(rows[0][columnIndex], rows[rowIndex][0])] = rows[rowIndex][columnIndex]

    # ROOM REQUIREMENTS
    #     -Echelons are based on set up start times
    requirement_df = xl.parse("Event Requirements")
    requirement_rows = requirement_df.values.tolist()

    event_room_list = []
    item_list = []
    event_dict = {}
    echelon_list = []
    echelon_count = 0
    for row in requirement_rows:
        if any(row):
            event = row[0]
            room = row[1]
            start_setup = row[2]
            req_end = row[5]
            commodity = row[6]
            quantity = row[7]

            if room not in event_room_list:
                event_room_list.append(room)

            if commodity not in item_list:
                item_list.append(commodity)

            if event not in event_dict:
                event_dict[event] = ([start_setup], [req_end])
            else:
                event_dict[event][0].append(start_setup)
                event_dict[event][1].append(req_end)

            if start_setup not in echelon_list:
                echelon_list.append(start_setup)
                echelon_count += 1
            if req_end not in echelon_list:
                echelon_list.append(req_end)
                echelon_count += 1

    echelon_list = sorted(echelon_list)
    echelon_dict = {i + 1: echelon_list[i] for i in range(len(echelon_list))}

    # COMMODITIES
    items_df = xl.parse("Commodities")
    items_rows = items_df.values.tolist()

    item_dict = {}
    priority_dict = {}
    for row in items_rows:
        commodity = row[0]
        com_per_parcel = float(row[1])
        vol_per_parcel = float(row[2])
        priority = int(row[3])
        if any(row):
            if commodity in item_list:
                item_dict[commodity] = (com_per_parcel, vol_per_parcel)
                priority_dict[commodity] = priority

    priority_list = sorted(priority_dict, key=lambda k: priority_dict[k])

    # event_room_list = [ x[0] for x in list(inventory_dict.keys())]

    return (inventory_dict, total_inventory_dict, storage_cap_dict, cost_dict, requirement_rows, event_room_list, event_dict, echelon_dict, item_dict, priority_list)

def outer(event_dict, requirement_rows, total_inventory_dict):
    event_times_dict = {}
    for event in event_dict:
        earliest_setup = min(event_dict[event][0])
        latest_end = max(event_dict[event][1])
        event_times_dict[event] = (earliest_setup, latest_end)
    sorted_events = sorted(event_times_dict, key=lambda k: event_times_dict[k][0])
    # print(sorted_events)
    # print(event_times_dict)

    i = 0
    super_event_dict = {}
    # print(len(sorted_events))
    while i < (len(sorted_events)):
        event = sorted_events[i]
        # print(i)
        # print(event)
        if i == (len(sorted_events) - 1):
            super_event_dict[event] = event_times_dict[event]
            i += 1
        else:
            next_event = sorted_events[i + 1]
            if event_times_dict[event][1] < event_times_dict[next_event][0]:
                super_event_dict[event] = event_times_dict[event]
                i += 1
            else:
                conglomerate_name = event + ' and ' + next_event
                # super_event_dict[conglomerate_name] = (event_times_dict[event][0], max(event_times_dict[next_event][0], event_times_dict[next_event][1])) ## WHY MAX?
                super_event_dict[conglomerate_name] = (event_times_dict[event][0], event_times_dict[next_event][1])
                i += 2

    print(super_event_dict)

    i = 0
    echelon_dict = {}
    echelon_dict_reverse = {}
    event_requirement_dict = {}
    sorted_super_events = sorted(super_event_dict, key=lambda k: super_event_dict[k][0])
    while i < len(sorted_super_events):
        echelon_dict[i*2 + 1] = super_event_dict[sorted_super_events[i]][0]
        echelon_dict_reverse[super_event_dict[sorted_super_events[i]][0]] = i*2 + 1
        echelon_dict[i*2+2] = super_event_dict[sorted_super_events[i]][1]
        echelon_dict_reverse[super_event_dict[sorted_super_events[i]][1]] = i*2 + 1
        event_requirement_dict[sorted_super_events[i]] = [] #reqs for superevent
        i += 1

    for row in requirement_rows:
        i = 0
        while i < len(sorted_super_events) :
            start_setup = row[2]
            if start_setup >= echelon_dict[2*i + 1] and start_setup < echelon_dict[2*i + 2]:
                event_requirement_dict[sorted_super_events[i]].append(row)
                i = len(sorted_super_events)
            i += 1

    equip_counter = {com:0 for com in total_inventory_dict}
    requirement_dict  = {}
    for sEvent in event_requirement_dict:
        event_start = super_event_dict[sEvent][0]
        req_end = super_event_dict[sEvent][1]
        for requirement in event_requirement_dict[sEvent]:
            room = requirement[1]
            commodity = requirement[6]
            quantity = requirement[7]
            total_of_item = total_inventory_dict[commodity]
            counter = equip_counter[commodity]
            availible = total_of_item - counter
            qty = min(quantity, availible)
            if (room, echelon_dict_reverse[event_start]) in requirement_dict:
                if commodity in requirement_dict[(room, echelon_dict_reverse[event_start])]:
                    requirement_dict[(room, echelon_dict_reverse[event_start])][commodity] = max(qty, requirement_dict[(room, echelon_dict_reverse[event_start])][commodity])
                    equip_counter[commodity] -= requirement_dict[(room, echelon_dict_reverse[event_start])][commodity]
                    equip_counter[commodity] += max(qty, requirement_dict[(room, echelon_dict_reverse[event_start])][commodity])
                else:
                    requirement_dict[(room, echelon_dict_reverse[event_start])][commodity] = qty
                    equip_counter[commodity] += qty
            else:
                requirement_dict[(room, echelon_dict_reverse[event_start])] = {commodity: qty}
                equip_counter[commodity] += qty
    return echelon_dict, requirement_dict

def inner(start, end, requirement_rows):
    echelons = []
    for row in requirement_rows:
        event = row[0]
        room = row[1]
        start_setup = row[2]
        req_end = row[5]
        commodity = row[6]
        quantity = row[7]
        if start <= start_setup and start_setup < end:
            if start_setup not in echelons:
                echelons.append(start_setup)
            if req_end not in echelons:
                echelons.append(req_end)
    echelons = sorted(echelons)
    echelon_dict = {i+1 : echelons[i] for i in range(len(echelons))}

    requirement_dict = {}
    for row in requirement_rows:
        event = row[0]
        room = row[1]
        start_setup = row[2]
        req_end = row[5]
        commodity = row[6]
        quantity = row[7]
        i = 0
        while i < len(echelon_dict):
            if (start_setup <= echelons[i]) :
                if (echelons[i] < req_end):
                    if (room, i +1) in requirement_dict:
                            requirement_dict[(room, i +1)][commodity] = quantity
                    else:
                        requirement_dict[(room, i +1)] = {commodity: quantity}
                    i += 1
                else:
                    i = len(echelons)
            else:
                i += 1
    return echelon_dict, requirement_dict

def outerConstructor(echelon_dict, eventRoomList, item_dict, costDict, requirementDict, inventory_dict, total_inventory_dict, storage_cap_dict):
    # print("costdict:")
    # for (roomI, roomJ) in costDict:
    #     if roomI == 'A 301':
    #         print((roomI, roomJ))

    #Creates 4 sets of arcs:
    #   movement_arc_dict       All b to a, room to room movement arcs (decisions)
    #   storage_cap_arc_dict    All a to b storage arcs (decisions)
    #   event_req_arc_dict      All a to b event requirement arcs (givens)
    #   utility_arc_dict        All arcs originating at the s node or between t nodes (givens)

    storageRoomList = list(storage_cap_dict)
    active_room_list = eventRoomList + storageRoomList
    all_room_list = [x[0] for x in list(inventory_dict.keys())]
    #print(allRoomList)
    movement_arc_dict = {}
    storage_cap_arc_dict = {}
    event_req_arc_dict = {}
    utility_arc_dict = {}
    itemList = list(item_dict)

    #Create general set of arcs for all time echelons other than the first and
    #last
    sorted_echelon_list = sorted(list(echelon_dict.keys()))
    print(echelon_dict)
    print(sorted_echelon_list)
    for echelon in echelon_dict:
        for roomI in active_room_list:
            for roomJ in active_room_list:
                for ab in ["a", "b"]:
                    if ab == "a":
                        if roomI == roomJ:
                            if (roomI, echelon) in requirementDict:
                                if type(requirementDict[(roomI, echelon)]) == tuple:
                                    for requirement in requirementDict[(roomI, echelon)]:
                                        item, qty = requirement[0], requirement[1]
                                        event_req_arc_dict[((roomI, echelon, "a"),(roomJ, echelon, "b"), item)] = (qty, qty, 0)
                                elif type(requirementDict[(roomI, echelon)]) == dict:
                                    for item in requirementDict[(roomI, echelon)]:
                                        qty = requirementDict[(roomI, echelon)][item]
                                        event_req_arc_dict[((roomI, echelon, "a"),(roomJ, echelon, "b"), item)] = (qty, qty, 0)
                            if roomI in storageRoomList:
                                for item in itemList:
                                    # ub = int(storage_cap_dict[roomI] / item_dict[item][0]) ## MIGHT CHANGE WITH COMMODITY PAGE (UNITS/PARCEL)
                                    item_vol = item_dict[item][1] / item_dict[item][0]
                                    ub = int(storage_cap_dict[roomI] / item_vol)
                                    storage_cap_arc_dict[((roomI, echelon, "a"),(roomJ, echelon, "b"), item)] = (0, ub, 0)
                    if ab == "b":
                        if echelon != len(echelon_dict.keys()):
                            for item in itemList:
                                item_max = total_inventory_dict[item]
                                # item_cost = costDict[(roomI, roomJ)]
                                item_cost = costDict[(roomI, roomJ)]/item_dict[item][0]
                                movement_arc_dict[((roomI, echelon, "b"), (roomJ, echelon + 1, "a"), item)] = (0, item_max, item_cost)

    # for roomI in all_room_list:
    #     for roomJ in active_room_list:
    #         for item in itemList:
    #             movement_arc_dict[((roomI, 0, "b"), (roomJ, 1, "a"), item)] = (0, item_max, item_cost)

    #Create set of arcs for the last movement opportunity

    for room in active_room_list:
        for item in itemList:
            item_max = total_inventory_dict[item]
            movement_arc_dict[((room, (len(echelon_dict.keys())), "b"), ("t", (len(echelon_dict.keys()) + 1), "a"), item)] = (0, item_max, 0)

    #Create set of arcs for initial starting inventory

    for room in all_room_list:
        for item in itemList:
            utility_arc_dict[(("s", 0, "a"), (room, 0, "b"), item)] = (inventory_dict[(room, item)], inventory_dict[(room, item)], 0)
            # else:
            #     utility_arc_dict[(("s", 0, "a"), (room, 0, "b"), item)] = (0, 0, 0)

    #Create set of movement arcs for the first movement period

    for roomI in all_room_list:
        for roomJ in active_room_list:
            for item in itemList:
                item_max = total_inventory_dict[item]
                # item_cost = costDict[(roomI, roomJ)]
                item_cost = costDict[(roomI, roomJ)]/item_dict[item][0]
                movement_arc_dict[((roomI, 0, "b"), (roomJ, 1, "a"), item)] = (0, item_max, item_cost)

    rooms = len(active_room_list)
    for item in itemList:
        utility_arc_dict[(("t", (len(echelon_dict.keys()) + 1), "a"), ("t", (len(echelon_dict.keys()) + 1), "b"), item)] = (total_inventory_dict[item], total_inventory_dict[item], 0)

    return movement_arc_dict, storage_cap_arc_dict, event_req_arc_dict, utility_arc_dict, active_room_list

def innerConstructor(echelon_dict, eventRoomList, item_dict, costDict, requirementDict, total_inventory_dict, storage_cap_dict, start_state, end_state, inventory_dict):

    #Creates 4 sets of arcs:
    #   movement_arc_dict       All b to a, room to room movement arcs (decisions)
    #   storage_cap_arc_dict    All a to b storage arcs (decisions)
    #   event_req_arc_dict      All a to b event requirement arcs (givens)
    #   utility_arc_dict        All arcs originating at the s node or between t nodes (givens)

    storageRoomList = list(storage_cap_dict)
    active_room_list = eventRoomList + storageRoomList
    all_room_list = [x[0] for x in list(inventory_dict.keys())]
    #print(allRoomList)
    movement_arc_dict = {}
    storage_cap_arc_dict = {}
    event_req_arc_dict = {}
    utility_arc_dict = {}
    itemList = list(item_dict)

    #Create general set of arcs for all time echelons other than the first and
    #last
    for echelon in echelon_dict:
        for roomI in active_room_list:
            for roomJ in active_room_list:
                for ab in ["a", "b"]:
                    if ab == "a":
                        if roomI == roomJ:
                            if (roomI, echelon) in requirementDict:
                                if type(requirementDict[(roomI, echelon)]) == tuple:
                                    for requirement in requirementDict[(roomI, echelon)]:
                                        item, qty = requirement[0], requirement[1]
                                        event_req_arc_dict[((roomI, echelon, "a"),(roomJ, echelon, "b"), item)] = (qty, qty, 0)
                                elif type(requirementDict[(roomI, echelon)]) == dict:
                                    for item in requirementDict[(roomI, echelon)]:
                                        qty = requirementDict[(roomI, echelon)][item]
                                        event_req_arc_dict[((roomI, echelon, "a"),(roomJ, echelon, "b"), item)] = (qty, qty, 0)
                            if roomI in storageRoomList:
                                for item in itemList:
                                    # ub = int(storage_cap_dict[roomI] / item_dict[item][0]) ## MIGHT CHANGE WITH COMMODITY PAGE (UNITS/PARCEL)
                                    item_vol = item_dict[item][1] / item_dict[item][0]
                                    ub = int(storage_cap_dict[roomI] / item_vol)
                                    storage_cap_arc_dict[((roomI, echelon, "a"),(roomJ, echelon, "b"), item)] = (0, ub, 0)
                    if ab == "b":
                        for item in itemList:
                            item_max = total_inventory_dict[item]
                            # item_cost = costDict[(roomI, roomJ)]
                            item_cost = costDict[(roomI, roomJ)]/item_dict[item][0]
                            movement_arc_dict[((roomI, echelon, "b"), (roomJ, echelon + 1, "a"), item)] = (0, item_max, item_cost)

    last_echelon = sorted(echelon_dict)[-1]
    print(sorted(echelon_dict))
    for roomI in active_room_list:
        for roomJ in active_room_list:
            for item in itemList:
                item_max = total_inventory_dict[item]
                # item_cost = costDict[(roomI, roomJ)]
                item_cost = costDict[(roomI, roomJ)]/item_dict[item][0]
                movement_arc_dict[((roomI, last_echelon, "b"), (roomJ, last_echelon + 1, "a"), item)] = (0, item_max, item_cost)
    #Create set of arcs for inital starting conditions from s node to each room

    # for room in allRoomList:
    #     for item in itemList:
    #         item_max = total_inventory_dict[item]
    #         movement_arc_dict[((room, (len(echelon_dict.keys())), "b"), ("t", (len(echelon_dict.keys()) + 1), "a"), item)] = (0, item_max, 0)
    #         if room in storageRoomList:
    #             utility_arc_dict[(("s", 0, "a"), (room, 0, "b"), item)] = (inventory_dict[(room, item)], inventory_dict[(room, item)], 0)
    #         else:
    #             utility_arc_dict[(("s", 0, "a"), (room, 0, "b"), item)] = (0, 0, 0)

    for room in start_state:
        for item in start_state[room]:
            item_max = total_inventory_dict[item]
            utility_arc_dict[(("s", 0, "a"), (room, 0, "b"), item)] = (start_state[room][item], start_state[room][item], 0)

    #Create set of movement arcs for the first movement period

    for roomI in active_room_list:
        for roomJ in active_room_list:
            for item in itemList:
                item_max = total_inventory_dict[item]
                # item_cost = costDict[(roomI, roomJ)]
                item_cost = costDict[(roomI, roomJ)]/item_dict[item][0]
                movement_arc_dict[((roomI, 0, "b"), (roomJ, 1, "a"), item)] = (0, item_max, item_cost)

    print("endstats")
    print(end_state)
    for room in end_state:
        for item in start_state[room]:
            utility_arc_dict[((room, (len(echelon_dict.keys())), "a"), ("t", (len(echelon_dict.keys())), "b"), item)] = (end_state[room][item], end_state[room][item], 0)

    # rooms = len(allRoomList)
    # for item in itemList:
    #     utility_arc_dict[(("t", (len(echelon_dict.keys()) + 1), "a"), ("t", (len(echelon_dict.keys()) + 1), "b"), item)] = (total_inventory_dict[item], total_inventory_dict[item], 0)

    return movement_arc_dict, storage_cap_arc_dict, event_req_arc_dict, utility_arc_dict, active_room_list

def dataFramer(arcDict):
    arcList = []
    for arc in arcDict.keys():
        arcList.append([arc[0][0], arc[0][1], arc[0][2], arc[1][0], arc[1][1],
            arc[1][2], arc[2], arcDict[arc][0], arcDict[arc][1], arcDict[arc][2]])

    df = pd.DataFrame(arcList, columns=["Xi", "Yi", "Zi", "Xj", "Yj", "Zj", "Item", "Lij", "Uij", "Cij"])
    return df

def excelOutputWriter(solution, echelon_dict):
    arcList = []
    arcList.append(["Time", "echelon", "From Room", "To Room", "Commodity", "Amount"])
    for x in sorted(sorted(sorted(solution, key=lambda k: k[1][0]), key=lambda k: k[0][0]), key=lambda k: k[0][1]): ## Sort First by time, then by room???
        tail, head, commodity = x
        # print(x)
        # if solution[x] > 0:
            # if tail[0] != head[0]:
        print(str(x) + ": " + str(solution[x]))
            # print(echelon_dict[(tail[1])])
        if (head[1]) in echelon_dict:
            arcList.append([echelon_dict[(head[1])], head[1], tail[0], head[0], commodity, solution[x]])

    book = load_workbook(excel_filename)
    sheet_names = book.get_sheet_names()
    if 'Schedule' in sheet_names:
        sheet = book.get_sheet_by_name('Schedule')
        book.remove_sheet(sheet)
        book.save(excel_filename)
    df = pd.DataFrame(arcList)
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name="Schedule", index=False, index_label=False, header=False)
    writer.save()


def main(args):
    #sup()
    eventReader()

def sup():
    (inventory_dict, total_inventory_dict, storage_cap_dict, cost_dict, requirement_rows, event_room_list, event_dict, echelon_dict, item_dict, priority_list) = excelReader()
    echelon_dict, requirement_dict = outer(event_dict, requirement_rows, total_inventory_dict)
    movement_arc_dict, storage_cap_arc_dict, event_req_arc_dict, utility_arc_dict, allRoomList = outerConstructor(echelon_dict, event_room_list, item_dict, cost_dict, requirement_dict, inventory_dict, total_inventory_dict, storage_cap_dict)
    movement_arc_df = dataFramer(movement_arc_dict)
    storage_cap_arc_df = dataFramer(storage_cap_arc_dict)
    event_req_arc_df = dataFramer(event_req_arc_dict)
    utility_arc_df = dataFramer(utility_arc_dict)
    df_dict = {'movement': movement_arc_df, 'storage': storage_cap_arc_df, 'event': event_req_arc_df, 'utility': utility_arc_df, 'storage_rooms': storage_cap_dict, 'commodities': item_dict}
    print("\a")

    return(df_dict, cost_dict, priority_list, echelon_dict, event_room_list, item_dict, requirement_rows, total_inventory_dict, storage_cap_dict, inventory_dict)


def innerMCNF(start, end, start_state, end_state, event_room_list, item_dict, cost_dict, requirement_rows, total_inventory_dict, storage_cap_dict, inventory_dict):
    echelon_dict, requirement_dict = inner(start, end, requirement_rows)
    movement_arc_dict, storage_cap_arc_dict, event_req_arc_dict, utility_arc_dict, allRoomList = innerConstructor(echelon_dict, event_room_list, item_dict, cost_dict, requirement_dict, total_inventory_dict, storage_cap_dict, start_state, end_state, inventory_dict)
    movement_arc_df = dataFramer(movement_arc_dict)
    storage_cap_arc_df = dataFramer(storage_cap_arc_dict)
    event_req_arc_df = dataFramer(event_req_arc_dict)
    utility_arc_df = dataFramer(utility_arc_dict)
    df_dict = {'movement': movement_arc_df, 'storage': storage_cap_arc_df, 'event': event_req_arc_df, 'utility': utility_arc_df, 'storage_rooms': storage_cap_dict, 'commodities': item_dict}
    print("\a")
    return(df_dict, cost_dict, echelon_dict, event_room_list, item_dict, requirement_rows, total_inventory_dict, storage_cap_dict)

if __name__ == '__main__':
    import sys
    main(sys.argv)





# def arcDictWriter(arcDict, filename):
#     #Writes to csv file
#     arcList = []
#     for arc in arcDict.keys():
#         arcList.append([arc[0][0], arc[0][1], arc[0][2], arc[1][0], arc[1][1],
#             arc[1][2], arc[2], arcDict[arc][0], arcDict[arc][1], arcDict[arc][2]])
#     with open(filename, "w") as f:
#         writer = csv.writer(f)
#         writer.writerow(["Xi", "Yi", "Zi", "Xj", "Yj", "Zj", "Item", "Lij", "Uij", "Cij"])
#         writer.writerows(arcList)

